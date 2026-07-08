"""
Import/Export utility for employee data.

Supports CSV and XLSX formats for both import and export.
"""

import csv
import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


TEMPLATE_HEADERS = [
    "full_name",
    "employee_number",
    "email",
    "phone_number",
    "job_title",
    "department",
    "salary",
    "hire_date",
    "status",
]

TEMPLATE_EXAMPLE = [
    "John Doe",
    "EMP-001",
    "john@company.com",
    "1012345678",
    "Software Engineer",
    "engineering",
    "50000.00",
    "2026-01-15",
    "active",
]

FIELD_LABELS = {
    "full_name": "Full Name",
    "employee_number": "Employee #",
    "email": "Email",
    "phone_number": "Phone Number",
    "job_title": "Job Title",
    "department": "Department",
    "salary": "Salary",
    "hire_date": "Hire Date",
    "status": "Status",
}

REQUIRED_FIELDS = {"full_name", "employee_number", "email", "salary", "hire_date"}


# ── XLSX generation helpers ───────────────────────────────────────────────────

def _style_header(ws):
    """Apply header styling to the first row."""
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for col_idx in range(1, len(TEMPLATE_HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.auto_filter.ref = ws.dimensions


def _write_rows_to_xlsx(rows, headers):
    """Write rows (list of dicts) to an XLSX workbook and return bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    _style_header(ws)

    for row_idx, row in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(header))

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _write_rows_to_csv(rows, headers):
    """Write rows (list of dicts) to CSV and return bytes."""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=headers)
    writer.writeheader()
    for row in rows:
        writer.writerow({h: row.get(h, "") for h in headers})
    return buf.getvalue().encode("utf-8-sig")


# ── Public API ────────────────────────────────────────────────────────────────


def generate_import_template() -> bytes:
    """Generate an XLSX template file with headers and one example row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"

    for col_idx, header in enumerate(TEMPLATE_HEADERS, 1):
        ws.cell(row=1, column=col_idx, value=header)
    _style_header(ws)

    for col_idx, value in enumerate(TEMPLATE_EXAMPLE, 1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.font = Font(color="999999", italic=True)

    for col_idx in range(1, len(TEMPLATE_HEADERS) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def parse_import_file(content: bytes, filename: str) -> list[dict[str, Any]]:
    """
    Parse uploaded file content into a list of row dicts.
    Supports both .xlsx and .csv.
    """
    filename_lower = filename.lower()

    if filename_lower.endswith(".xlsx"):
        return _parse_xlsx(content)
    elif filename_lower.endswith(".csv"):
        return _parse_csv(content)
    else:
        raise ValueError("Unsupported file format. Please upload .xlsx or .csv.")


def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    if ws is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    if not rows or not rows[0]:
        return []

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    results = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
        row_dict = {}
        for idx, value in enumerate(row):
            if idx < len(headers) and headers[idx]:
                cell = value
                if isinstance(cell, datetime):
                    cell = cell.date().isoformat()
                elif isinstance(cell, date):
                    cell = cell.isoformat()
                elif isinstance(cell, (int, float, Decimal)):
                    cell = str(cell)
                elif cell is None:
                    cell = ""
                else:
                    cell = str(cell).strip()
                row_dict[headers[idx]] = cell
        results.append(row_dict)
    return results


def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    headers = [h.strip().lower() for h in reader.fieldnames] if reader.fieldnames else []
    results = []
    for row in reader:
        row_dict = {}
        for header in headers:
            val = row.get(header, "").strip()
            if val == "":
                val = ""
            row_dict[header] = val
        results.append(row_dict)
    return results


def validate_import_rows(
    rows: list[dict[str, Any]],
    existing_employees: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    """
    Validate a list of parsed rows against business rules.

    Returns (valid_rows, errors) where errors is a list of
    {row, reason} dicts.
    """
    valid = []
    errors = []

    existing_numbers = {e["employee_number"] for e in existing_employees if e.get("employee_number")}
    existing_emails = {e["email"] for e in existing_employees if e.get("email")}
    seen_numbers = set()
    seen_emails = set()

    for idx, row in enumerate(rows, 1):
        reasons = []

        for field in REQUIRED_FIELDS:
            if not row.get(field):
                reasons.append(f"'{FIELD_LABELS.get(field, field)}' is required.")

        full_name = row.get("full_name", "")
        if len(full_name) < 2 or len(full_name) > 160:
            reasons.append("'Full Name' must be 2-160 characters.")

        emp_number = row.get("employee_number", "")
        if emp_number and (len(emp_number) < 1 or len(emp_number) > 30):
            reasons.append("'Employee #' must be 1-30 characters.")

        email = row.get("email", "")
        if email and ("@" not in email or len(email) > 255):
            reasons.append("Enter a valid email address.")

        salary_str = row.get("salary", "")
        salary = None
        if salary_str:
            try:
                salary = Decimal(str(salary_str))
                if salary < 0:
                    reasons.append("'Salary' must be 0 or greater.")
            except Exception:
                reasons.append("'Salary' must be a valid number.")

        hire_date_str = row.get("hire_date", "")
        if hire_date_str:
            try:
                from datetime import date as date_type
                date_type.fromisoformat(hire_date_str)
            except (ValueError, TypeError):
                reasons.append("'Hire Date' must be YYYY-MM-DD format.")

        status = row.get("status", "")
        if status and status not in ("active", "resigned"):
            reasons.append("'Status' must be 'active' or 'resigned'.")

        if emp_number:
            if emp_number in existing_numbers or emp_number in seen_numbers:
                reasons.append(f"Employee # '{emp_number}' already exists.")
            seen_numbers.add(emp_number)

        if email:
            if email in existing_emails or email in seen_emails:
                reasons.append(f"Email '{email}' already exists.")
            seen_emails.add(email)

        if reasons:
            errors.append({"row": idx, "reasons": "; ".join(reasons)})
        else:
            row_dict = {
                "full_name": full_name,
                "employee_number": emp_number,
                "email": email,
                "phone_number": row.get("phone_number") or None,
                "job_title": row.get("job_title") or None,
                "department": row.get("department") or None,
                "salary": salary if salary else Decimal("0"),
                "hire_date": hire_date_str,
                "status": status if status else "active",
            }
            valid.append(row_dict)

    return valid, errors


def generate_export(
    employees: list[dict[str, Any]],
    file_format: str,
    headers: list[str] | None = None,
) -> bytes:
    """
    Generate an export file (XLSX or CSV) from a list of employee dicts.

    The employees should contain at least the fields listed in TEMPLATE_HEADERS.
    """
    if headers is None:
        headers = TEMPLATE_HEADERS

    rows = []
    for emp in employees:
        row = {}
        for h in headers:
            val = emp.get(h)
            if isinstance(val, (date, datetime)):
                val = val.isoformat() if hasattr(val, "isoformat") else str(val)
            elif isinstance(val, Decimal):
                val = float(val)
            elif val is None:
                val = ""
            row[h] = val
        rows.append(row)

    if file_format == "csv":
        return _write_rows_to_csv(rows, headers)
    else:
        return _write_rows_to_xlsx(rows, headers)
